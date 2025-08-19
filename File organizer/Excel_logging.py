from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime


def log_file_move(filename, destination, logfile="log.xlsx"):
    log_path=Path(logfile)

    if log_path.exists():
        wb=load_workbook(logfile)
        ws=wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Filename", "Destination", "Date"])

    ws.append([filename, destination,datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(logfile)
    print(f"Logged {filename} -> {destination}")

